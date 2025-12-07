import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------
# Define the duration buckets
# -----------------------------
BUCKETS = [
    (0, 1),
    (1, 5),
    (5, 10),
    (10, 15),
    (15, 20),
    (20, 25),
    (25, 30),
    (30, float('inf'))  # 30+ seconds
]

def assign_bucket(duration):
    """Return the bucket label for a given duration."""
    for low, high in BUCKETS:
        if low <= duration < high:
            if high == float('inf'):
                return f"[{low}+)"
            return f"[{low}, {high})"
    return "UNBUCKETED"


def get_bucket_labels():
    """Get all bucket labels in order."""
    labels = []
    for low, high in BUCKETS:
        if high == float('inf'):
            labels.append(f"[{low}+)")
        else:
            labels.append(f"[{low}, {high})")
    return labels


# -----------------------------------------
# Recursive extraction of durations
# -----------------------------------------
def walk(node, folder_path, folder_stats, overall_buckets):
    """
    Recursively walk the JSON structure and collect statistics.
    folder_stats: dict with folder_path -> {buckets, total_files, total_duration}
    """
    # Initialize folder stats if not exists
    if folder_path not in folder_stats:
        folder_stats[folder_path] = {
            'buckets': {label: 0 for label in get_bucket_labels()},
            'total_files': 0,
            'total_duration': 0.0
        }

    # Process FLAC files in current directory
    if "flac_files" in node:
        for f in node["flac_files"]:
            dur = f["duration_seconds"]
            bucket = assign_bucket(dur)
            
            # Update folder stats
            folder_stats[folder_path]['buckets'][bucket] += 1
            folder_stats[folder_path]['total_files'] += 1
            folder_stats[folder_path]['total_duration'] += dur
            
            # Update overall buckets
            overall_buckets[bucket] += 1

    # Recurse into subdirectories
    for sub in node.get("subdirectories", []):
        sub_path = os.path.join(folder_path, sub["name"])
        walk(sub, sub_path, folder_stats, overall_buckets)


# -----------------------------
# Main execution function
# -----------------------------
def process_dataset(json_data, dataset_name):
    """
    Process a dataset and return statistics.
    
    Returns:
        tuple: (overall_buckets, folder_stats, total_files)
    """
    bucket_labels = get_bucket_labels()
    overall_buckets = {label: 0 for label in bucket_labels}
    folder_stats = {}

    # Start recursive walking
    root_path = json_data["name"]
    walk(json_data, root_path, folder_stats, overall_buckets)

    # Calculate overall totals
    total_files = sum(overall_buckets.values())
    
    # Print summary to console
    print(f"\n{'='*60}")
    print(f"Dataset: {dataset_name}")
    print(f"{'='*60}")
    print(f"Total files: {total_files:,}")
    print(f"Total folders: {len(folder_stats)}")
    print(f"\nDuration Distribution:")
    for label in bucket_labels:
        count = overall_buckets[label]
        pct = (count / total_files * 100) if total_files > 0 else 0
        print(f"  {label:15s}: {count:6,} files ({pct:5.2f}%)")
    
    return overall_buckets, folder_stats, total_files


def write_to_excel(datasets_data, output_file="audio_distribution_analysis.xlsx"):
    """
    Write all dataset statistics to a single Excel file with multiple sheets.
    
    Args:
        datasets_data: dict with dataset_name -> (overall_buckets, folder_stats, total_files)
        output_file: Output Excel file path
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    bucket_labels = get_bucket_labels()
    
    # Header styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for dataset_name, (overall_buckets, folder_stats, total_files) in datasets_data.items():
        # Sheet 1: Overall Distribution
        ws_overall = wb.create_sheet(f"{dataset_name}_overall")
        ws_overall.append(["Bucket", "Count", "Percentage"])
        
        # Style header
        for cell in ws_overall[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data
        for label in bucket_labels:
            count = overall_buckets[label]
            pct = (count / total_files * 100) if total_files > 0 else 0
            ws_overall.append([label, count, f"{pct:.2f}%"])
        
        ws_overall.append(["TOTAL", total_files, "100.00%"])
        
        # Make last row bold
        for cell in ws_overall[ws_overall.max_row]:
            cell.font = Font(bold=True)
        
        # Auto-size columns
        for col in range(1, 4):
            ws_overall.column_dimensions[get_column_letter(col)].width = 15
        
        # Sheet 2: Folder Summary
        ws_summary = wb.create_sheet(f"{dataset_name}_summary")
        ws_summary.append([
            "Folder", "Total Files", "Total Duration (sec)", 
            "Avg Duration (sec)", "Duration (hours)"
        ])
        
        # Style header
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data
        for folder in sorted(folder_stats.keys()):
            stats = folder_stats[folder]
            avg_dur = stats['total_duration'] / stats['total_files'] if stats['total_files'] > 0 else 0
            dur_hours = stats['total_duration'] / 3600
            
            ws_summary.append([
                folder,
                stats['total_files'],
                round(stats['total_duration'], 2),
                round(avg_dur, 2),
                round(dur_hours, 2)
            ])
        
        # Auto-size columns
        ws_summary.column_dimensions['A'].width = 50
        for col in range(2, 6):
            ws_summary.column_dimensions[get_column_letter(col)].width = 18
        
        # Sheet 3: Folder-wise Distribution
        ws_folder = wb.create_sheet(f"{dataset_name}_folderwise")
        ws_folder.append(["Folder", "Total Files", "Total Duration (sec)"] + bucket_labels)
        
        # Style header
        for cell in ws_folder[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data
        for folder in sorted(folder_stats.keys()):
            stats = folder_stats[folder]
            row = [
                folder,
                stats['total_files'],
                round(stats['total_duration'], 2)
            ] + [stats['buckets'][label] for label in bucket_labels]
            ws_folder.append(row)
        
        # Auto-size columns
        ws_folder.column_dimensions['A'].width = 50
        for col in range(2, len(bucket_labels) + 4):
            ws_folder.column_dimensions[get_column_letter(col)].width = 12
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ All data written to: {output_file}")
    print(f"  Sheets created: {len(wb.sheetnames)}")
    for sheet_name in wb.sheetnames:
        print(f"    - {sheet_name}")


# -----------------------------
# Example usage
# -----------------------------
if __name__ == "__main__":
    # Process both datasets
    datasets = [
        ("create_structure.json", "create"),
        ("collect_structure.json", "collect")
    ]
    
    datasets_data = {}
    
    for json_file, dataset_name in datasets:
        if os.path.exists(json_file):
            print(f"\nProcessing {json_file}...")
            with open(json_file, "r", encoding="utf-8") as fp:
                data = json.load(fp)
            
            overall_buckets, folder_stats, total_files = process_dataset(data, dataset_name)
            datasets_data[dataset_name] = (overall_buckets, folder_stats, total_files)
        else:
            print(f"⚠ Warning: {json_file} not found, skipping...")
    
    # Write all data to single Excel file
    if datasets_data:
        write_to_excel(datasets_data)